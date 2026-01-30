#!/usr/bin/env python3
"""
Parse Chrome bookmarks HTML export and categorize into markdown files and Excel spreadsheet.
"""

import os
import re
from datetime import datetime
from pathlib import Path
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def convert_timestamp(unix_time):
    """Convert Unix timestamp to readable date."""
    if unix_time:
        try:
            return datetime.fromtimestamp(int(unix_time)).strftime('%Y-%m-%d')
        except (ValueError, OSError):
            return "Unknown"
    return "Unknown"


def categorize_bookmark(url, title):
    """Categorize bookmark based on URL and title patterns."""
    url_lower = url.lower() if url else ""
    title_lower = title.lower() if title else ""
    
    # Tour Operators
    if any(x in url_lower for x in ['getyourguide', 'withlocals', 'bokun', 'fhapi', 'booking.com', 'viator', 'toursByLocals']):
        return 'Tour Operators'
    if 'tour' in title_lower and any(x in url_lower for x in ['guide', 'operator', 'booking']):
        return 'Tour Operators'
    
    # Travel Blogs
    if any(x in url_lower for x in ['blog', 'secreta', 'theurgetowander', 'lisbonlux', 'cityguidelisbon', 'thegeographicalcure', 'exploredbymarta']):
        return 'Travel Blogs'
    if any(x in title_lower for x in ['blog', 'guide', 'travel', 'hidden gems', 'secret spots']):
        return 'Travel Blogs'
    
    # Event Sites
    if any(x in url_lower for x in ['agendalx', 'timeout', 'agenda', 'eventbrite', 'bandsintown', 'blueticket']):
        return 'Event Sites'
    if 'event' in title_lower or 'festa' in title_lower:
        return 'Event Sites'
    
    # Media/News
    if any(x in url_lower for x in ['theportugalnews', 'expresso', 'visao', 'nit.pt', 'news', 'media']):
        return 'Media/News'
    if any(x in title_lower for x in ['news', 'portugal', 'media']):
        return 'Media/News'
    
    # Food/Venues
    if any(x in url_lower for x in ['mesamarcada', 'musicbox', 'restaurant', 'bar', 'cafe', 'food', 'lojas']):
        return 'Food/Venues'
    if any(x in title_lower for x in ['restaurant', 'bar', 'cafe', 'food', 'comida', 'bebida']):
        return 'Food/Venues'
    
    # Government/Official
    if any(x in url_lower for x in ['visitlisboa', 'egeac', 'turismo', 'official', 'governo']):
        return 'Government/Official'
    if 'visit' in title_lower or 'official' in title_lower:
        return 'Government/Official'
    
    # Transportation
    if any(x in url_lower for x in ['rede-expressos', 'bus', 'transport', 'ride', 'lisboaride']):
        return 'Transportation'
    if 'transport' in title_lower or 'bus' in title_lower:
        return 'Transportation'
    
    # Research/Design
    if any(x in url_lower for x in ['behance', 'dribbble', 'freepik', 'vector', 'hellovector', 'design', 'template', 'inspiration']):
        return 'Research/Design'
    if any(x in title_lower for x in ['design', 'template', 'vector', 'inspiration', 'asset']):
        return 'Research/Design'
    
    return 'Other'


def parse_bookmarks(html_file):
    """Parse Chrome bookmarks HTML file."""
    with open(html_file, 'r', encoding='utf-8') as f:
        html = f.read()
    
    soup = BeautifulSoup(html, 'html.parser')
    bookmarks = []
    
    for link in soup.find_all('a'):
        url = link.get('href', '')
        title = link.get_text(strip=True)
        date_added = link.get('add_date', '')
        
        # Skip empty entries
        if not url or not title:
            continue
        
        bookmark = {
            'url': url,
            'title': title,
            'date_added': convert_timestamp(date_added),
            'category': categorize_bookmark(url, title),
            'subcategory': '',
            'priority': '',
            'notes': '',
            'status': ''
        }
        bookmarks.append(bookmark)
    
    return bookmarks


def create_markdown_files(bookmarks, output_dir):
    """Create categorized markdown files."""
    categories = {}
    
    # Group bookmarks by category
    for bookmark in bookmarks:
        cat = bookmark['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(bookmark)
    
    # Create markdown files
    for category, items in sorted(categories.items()):
        # Sort by date (newest first)
        items.sort(key=lambda x: x['date_added'], reverse=True)
        
        filename = category.lower().replace('/', '-').replace(' ', '-') + '.md'
        filepath = os.path.join(output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# {category} - Chrome Bookmarks\n\n")
            f.write(f"**Total Bookmarks:** {len(items)}\n")
            f.write(f"**Last Updated:** {datetime.now().strftime('%b %d, %Y')}\n\n")
            f.write("---\n\n")
            
            for item in items:
                f.write(f"### [{item['title']}]({item['url']})\n")
                f.write(f"**Added:** {item['date_added']}\n\n")
        
        print(f"✓ Created: {filename} ({len(items)} bookmarks)")


def create_excel_file(bookmarks, output_file):
    """Create Excel spreadsheet with all bookmarks."""
    # Sort by category, then date (newest first)
    bookmarks.sort(key=lambda x: (x['category'], x['date_added']), reverse=(False, True))
    
    df = pd.DataFrame(bookmarks)
    
    # Reorder columns
    df = df[['URL', 'Title', 'Category', 'Subcategory', 'Date Added', 'Priority', 'Notes', 'Status']]
    
    # Create Excel file
    df.to_excel(output_file, index=False, sheet_name='Bookmarks')
    
    # Format Excel
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add autofilter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(df.columns))}{len(df) + 1}'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 10
    
    # Format header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    wb.save(output_file)
    print(f"✓ Created: {os.path.basename(output_file)} ({len(df)} rows)")


def create_summary_report(bookmarks, output_file):
    """Create summary report."""
    categories = {}
    for bookmark in bookmarks:
        cat = bookmark['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(bookmark)
    
    # Date distribution
    today = datetime.now()
    last_30 = sum(1 for b in bookmarks if b['date_added'] != 'Unknown' and 
                  (today - datetime.strptime(b['date_added'], '%Y-%m-%d')).days <= 30)
    last_90 = sum(1 for b in bookmarks if b['date_added'] != 'Unknown' and 
                  (today - datetime.strptime(b['date_added'], '%Y-%m-%d')).days <= 90)
    last_year = sum(1 for b in bookmarks if b['date_added'] != 'Unknown' and 
                    (today - datetime.strptime(b['date_added'], '%Y-%m-%d')).days <= 365)
    older = len(bookmarks) - last_year
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("# Chrome Links Analysis Summary\n\n")
        f.write(f"**Parsed:** {datetime.now().strftime('%b %d, %Y at %H:%M UTC')}\n")
        f.write(f"**Total Bookmarks:** {len(bookmarks)}\n")
        f.write(f"**Categories:** {len(categories)}\n\n")
        f.write("---\n\n")
        
        # Category breakdown
        f.write("## Category Breakdown\n\n")
        f.write("| Category | Count | Top 3 Sites |\n")
        f.write("|----------|-------|-------------|\n")
        
        for cat in sorted(categories.keys()):
            items = categories[cat]
            top_3 = ', '.join([item['title'][:25] for item in items[:3]])
            f.write(f"| {cat} | {len(items)} | {top_3} |\n")
        
        f.write("\n---\n\n")
        
        # Key findings
        f.write("## Key Findings\n\n")
        
        if 'Tour Operators' in categories:
            f.write(f"- **Tour Operators:** {len(categories['Tour Operators'])} bookmarks\n")
        
        if 'Travel Blogs' in categories:
            f.write(f"- **Travel Blogs:** {len(categories['Travel Blogs'])} bookmarks\n")
        
        if 'Event Sites' in categories:
            f.write(f"- **Event Sites:** {len(categories['Event Sites'])} bookmarks\n")
        
        f.write("\n---\n\n")
        
        # Date distribution
        f.write("## Date Distribution\n\n")
        f.write(f"- Last 30 days: {last_30} bookmarks\n")
        f.write(f"- Last 90 days: {last_90} bookmarks\n")
        f.write(f"- Last year: {last_year} bookmarks\n")
        f.write(f"- Older: {older} bookmarks\n")
        f.write(f"\n---\n\n")
        
        f.write("## Recommended Actions\n\n")
        f.write("1. Review Tour Operators for partnership opportunities\n")
        f.write("2. Extract content from Travel Blogs for inspiration\n")
        f.write("3. Monitor Event Sites for activity updates\n")
        f.write("4. Organize research/design assets for reuse\n")
    
    print(f"✓ Created: {os.path.basename(output_file)}")


def main():
    """Main execution."""
    # Define paths
    html_file = r"d:\Backup and Documents\W\Windsurf\## LBPF links\bookmarks_10_26_25.html"
    output_base = r"d:\Backup and Documents\W\Windsurf\lisboaporfavor-bootstrap\data\chrome-links"
    categorized_dir = os.path.join(output_base, 'categorized')
    
    # Create directories
    Path(categorized_dir).mkdir(parents=True, exist_ok=True)
    
    print("🔍 Parsing Chrome bookmarks...")
    bookmarks = parse_bookmarks(html_file)
    print(f"✓ Found {len(bookmarks)} bookmarks\n")
    
    # Rename columns for consistency
    for b in bookmarks:
        b['URL'] = b.pop('url')
        b['Title'] = b.pop('title')
        b['Date Added'] = b.pop('date_added')
        b['Category'] = b.pop('category')
        b['Subcategory'] = b.pop('subcategory')
        b['Priority'] = b.pop('priority')
        b['Notes'] = b.pop('notes')
        b['Status'] = b.pop('status')
    
    print("📝 Creating markdown files...")
    create_markdown_files(bookmarks, categorized_dir)
    
    print("\n📊 Creating Excel spreadsheet...")
    excel_file = os.path.join(output_base, 'chrome-links-full.xlsx')
    create_excel_file(bookmarks, excel_file)
    
    print("\n📋 Creating summary report...")
    summary_file = os.path.join(output_base, 'summary.md')
    create_summary_report(bookmarks, summary_file)
    
    print("\n✅ All files created successfully!")
    print(f"\n📁 Output directory: {output_base}")


if __name__ == '__main__':
    main()
